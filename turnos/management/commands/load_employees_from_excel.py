from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.db import transaction
from empleados.models import Empleado
from datetime import datetime
import openpyxl
import os


class Command(BaseCommand):
    help = 'Carga empleados desde un archivo Excel (PEsperanza.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Ruta del archivo Excel a cargar'
        )
        parser.add_argument(
            '--update',
            action='store_true',
            help='Actualizar empleados existentes en lugar de omitirlos'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='Hoja1',
            help='Nombre de la hoja a leer (default: Hoja1)'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        update_existing = options['update']
        sheet_name = options['sheet']
        
        # Verificar que el archivo existe
        if not os.path.exists(excel_file):
            raise CommandError(f'El archivo "{excel_file}" no existe.')
        
        self.stdout.write(self.style.SUCCESS(f'Cargando empleados desde: {excel_file}'))
        
        try:
            # Cargar el archivo Excel
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            # Intentar usar la hoja especificada, si no existe usar la primera
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
                self.stdout.write(
                    self.style.WARNING(
                        f'Hoja "{sheet_name}" no encontrada. Usando hoja activa: {sheet.title}'
                    )
                )
            
            # Leer encabezados (primera fila)
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip().lower() if cell.value else '')
            
            self.stdout.write(f'Encabezados encontrados: {headers}')
            
            # Mapear columnas esperadas
            col_map = self._map_columns(headers)
            
            # Estadísticas
            created = 0
            updated = 0
            skipped = 0
            errors = []
            
            # Procesar cada fila (desde la segunda en adelante)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    with transaction.atomic():
                        result = self._process_row(row, col_map, update_existing)
                        
                        if result == 'created':
                            created += 1
                        elif result == 'updated':
                            updated += 1
                        elif result == 'skipped':
                            skipped += 1
                            
                except Exception as e:
                    errors.append(f'Fila {row_num}: {str(e)}')
                    self.stdout.write(self.style.ERROR(f'Error en fila {row_num}: {str(e)}'))
            
            # Mostrar resumen
            self.stdout.write(self.style.SUCCESS('\n=== RESUMEN DE CARGA ==='))
            self.stdout.write(self.style.SUCCESS(f'Empleados creados: {created}'))
            if update_existing:
                self.stdout.write(self.style.SUCCESS(f'Empleados actualizados: {updated}'))
            self.stdout.write(self.style.WARNING(f'Empleados omitidos: {skipped}'))
            if errors:
                self.stdout.write(self.style.ERROR(f'Errores: {len(errors)}'))
                for error in errors[:10]:  # Mostrar primeros 10 errores
                    self.stdout.write(self.style.ERROR(f'  - {error}'))
            
            self.stdout.write(self.style.SUCCESS('\n¡Carga completada!'))
            
        except Exception as e:
            raise CommandError(f'Error al procesar el archivo: {str(e)}')
    
    def _map_columns(self, headers):
        """Mapea los encabezados del Excel a los campos del modelo"""
        # Posibles variaciones de nombres de columnas
        mapping = {
            'codigo_empleado': ['codigo', 'codigo_empleado', 'código', 'código empleado', 'id', 'employee_id'],
            'nombre': ['nombre', 'nombres', 'name', 'first_name'],
            'apellido': ['apellido', 'apellidos', 'last_name', 'surname'],
            'email': ['email', 'correo', 'e-mail', 'mail'],
            'departamento': ['departamento', 'depto', 'department', 'área', 'area'],
            'puesto': ['puesto', 'cargo', 'position', 'job_title'],
            'fecha_ingreso': ['fecha_ingreso', 'fecha ingreso', 'fecha de ingreso', 'hire_date', 'start_date'],
            'horas_semana': ['horas_semana', 'horas', 'hours', 'horas semanales'],
            'username': ['usuario', 'username', 'user'],
        }
        
        col_map = {}
        for field, variations in mapping.items():
            for i, header in enumerate(headers):
                if header in variations:
                    col_map[field] = i
                    break
        
        return col_map
    
    def _process_row(self, row, col_map, update_existing):
        """Procesa una fila del Excel y crea/actualiza el empleado"""
        
        # Obtener código de empleado (requerido)
        codigo_empleado = self._get_value(row, col_map, 'codigo_empleado')
        if not codigo_empleado:
            raise ValueError('Código de empleado es requerido')
        
        codigo_empleado = str(codigo_empleado).strip()
        
        # Verificar si el empleado ya existe
        try:
            empleado = Empleado.objects.get(codigo_empleado=codigo_empleado)
            if not update_existing:
                return 'skipped'
            is_new = False
        except Empleado.DoesNotExist:
            empleado = None
            is_new = True
        
        # Obtener datos del empleado
        nombre = self._get_value(row, col_map, 'nombre', '')
        apellido = self._get_value(row, col_map, 'apellido', '')
        email = self._get_value(row, col_map, 'email', '')
        departamento = self._get_value(row, col_map, 'departamento', 'General')
        puesto = self._get_value(row, col_map, 'puesto', '')
        horas_semana = self._get_value(row, col_map, 'horas_semana', 40)
        username = self._get_value(row, col_map, 'username', codigo_empleado)
        
        # Procesar fecha de ingreso
        fecha_ingreso_raw = self._get_value(row, col_map, 'fecha_ingreso')
        fecha_ingreso = self._parse_date(fecha_ingreso_raw)
        
        # Crear o actualizar usuario de Django
        if is_new:
            # Generar username único si ya existe
            base_username = str(username).strip()
            username_final = base_username
            counter = 1
            while User.objects.filter(username=username_final).exists():
                username_final = f"{base_username}{counter}"
                counter += 1
            
            user = User.objects.create_user(
                username=username_final,
                email=email if email else f'{username_final}@example.com',
                first_name=str(nombre).strip() if nombre else '',
                last_name=str(apellido).strip() if apellido else '',
                password='changeme123'  # Contraseña temporal
            )
            
            # Crear empleado
            empleado = Empleado.objects.create(
                user=user,
                codigo_empleado=codigo_empleado,
                departamento=str(departamento).strip(),
                puesto=str(puesto).strip(),
                fecha_ingreso=fecha_ingreso,
                horas_semana=int(horas_semana) if horas_semana else 40,
                activo=True
            )
            return 'created'
        else:
            # Actualizar empleado existente
            if nombre:
                empleado.user.first_name = str(nombre).strip()
            if apellido:
                empleado.user.last_name = str(apellido).strip()
            if email:
                empleado.user.email = email
            empleado.user.save()
            
            if departamento:
                empleado.departamento = str(departamento).strip()
            if puesto:
                empleado.puesto = str(puesto).strip()
            if fecha_ingreso:
                empleado.fecha_ingreso = fecha_ingreso
            if horas_semana:
                empleado.horas_semana = int(horas_semana)
            
            empleado.save()
            return 'updated'
    
    def _get_value(self, row, col_map, field, default=None):
        """Obtiene el valor de una celda de manera segura"""
        if field in col_map and col_map[field] < len(row):
            value = row[col_map[field]]
            return value if value is not None else default
        return default
    
    def _parse_date(self, date_value):
        """Parsea una fecha en varios formatos posibles"""
        if not date_value:
            return None
        
        # Si ya es un objeto datetime, convertir a date
        if isinstance(date_value, datetime):
            return date_value.date()
        
        # Si es string, intentar parsear
        if isinstance(date_value, str):
            date_value = date_value.strip()
            formats = [
                '%Y-%m-%d',
                '%d/%m/%Y',
                '%d-%m-%Y',
                '%m/%d/%Y',
                '%Y/%m/%d',
            ]
            
            for fmt in formats:
                try:
                    return datetime.strptime(date_value, fmt).date()
                except ValueError:
                    continue
        
        return None
