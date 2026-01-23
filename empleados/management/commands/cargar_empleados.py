"""
Comando para cargar empleados desde archivo Excel.
Uso: python manage.py cargar_empleados
     python manage.py cargar_empleados --archivo PEsperanza.xlsx
     python manage.py cargar_empleados --dry-run
"""

from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from empleados.models import Empleado
import openpyxl
import unicodedata
import re


class Command(BaseCommand):
    help = 'Carga empleados desde archivo Excel (PEsperanza.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--archivo',
            type=str,
            default='PEsperanza.xlsx',
            help='Ruta al archivo Excel (default: PEsperanza.xlsx)'
        )
        parser.add_argument(
            '--password',
            type=str,
            default='Pa$$$Word2026',
            help='Contraseña para todos los usuarios'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Mostrar lo que se haría sin crear registros'
        )
        parser.add_argument(
            '--update',
            action='store_true',
            help='Actualizar empleados existentes'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        password = options['password']
        dry_run = options['dry_run']
        update = options['update']

        self.stdout.write(f'Cargando empleados desde: {archivo}')

        if dry_run:
            self.stdout.write(self.style.WARNING('MODO DRY-RUN: No se crearán registros'))

        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error al abrir archivo: {e}'))
            return

        # Leer encabezados para mapear columnas
        headers = [cell.value for cell in ws[1]]
        self.stdout.write(f'Columnas encontradas: {headers}')

        # Mapeo de columnas (basado en PEsperanza.xlsx)
        col_map = {}
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                if header_lower in ['n°', 'no', 'numero', 'número', 'n']:
                    col_map['numero'] = idx
                elif header_lower in ['ofc', 'oficina', 'departamento', 'depto']:
                    col_map['departamento'] = idx
                elif header_lower in ['puesto', 'cargo', 'position']:
                    col_map['puesto'] = idx
                elif header_lower in ['titular', 'nombre', 'nombre completo', 'empleado']:
                    col_map['nombre'] = idx

        self.stdout.write(f'Mapeo de columnas: {col_map}')

        if 'numero' not in col_map or 'nombre' not in col_map:
            self.stdout.write(self.style.ERROR(
                'No se encontraron las columnas requeridas (N° y TITULAR/NOMBRE)'
            ))
            return

        creados = 0
        actualizados = 0
        existentes = 0
        errores = 0

        # Saltar la primera fila (encabezado)
        for row in ws.iter_rows(min_row=2, values_only=True):
            numero = row[col_map['numero']]
            nombre_completo = row[col_map['nombre']]
            departamento = row[col_map.get('departamento', 0)] if 'departamento' in col_map else 'General'
            puesto = row[col_map.get('puesto', 0)] if 'puesto' in col_map else ''

            if not numero or not nombre_completo:
                continue

            nombre_completo = str(nombre_completo).strip()
            departamento = str(departamento).strip() if departamento else 'General'
            puesto = str(puesto).strip() if puesto else ''

            # Generar código de empleado
            codigo = f'ESP{int(numero):03d}'

            # Generar username a partir del nombre
            username = self.generar_username(nombre_completo)

            # Separar nombre y apellidos (asumiendo formato: NOMBRE APELLIDO1 APELLIDO2)
            partes = nombre_completo.split()
            if len(partes) >= 3:
                # Primer nombre + resto son apellidos
                first_name = partes[0]
                last_name = ' '.join(partes[1:])
            elif len(partes) == 2:
                first_name = partes[0]
                last_name = partes[1]
            else:
                first_name = nombre_completo
                last_name = ''

            self.stdout.write(
                f'  {codigo}: {nombre_completo} ({puesto}) -> {username}'
            )

            if dry_run:
                continue

            try:
                # Verificar si el código de empleado ya existe
                empleado_existente = Empleado.objects.filter(codigo_empleado=codigo).first()

                if empleado_existente:
                    if update:
                        # Actualizar empleado existente
                        empleado_existente.departamento = departamento
                        empleado_existente.puesto = puesto
                        empleado_existente.user.first_name = first_name
                        empleado_existente.user.last_name = last_name
                        empleado_existente.user.save()
                        empleado_existente.save()
                        self.stdout.write(self.style.SUCCESS(f'    Actualizado: {empleado_existente}'))
                        actualizados += 1
                    else:
                        self.stdout.write(self.style.WARNING(f'    Empleado {codigo} ya existe'))
                        existentes += 1
                    continue

                # Verificar si el usuario ya existe (generar username único)
                username_final = username
                counter = 1
                while User.objects.filter(username=username_final).exists():
                    username_final = f'{username}{counter}'
                    counter += 1

                if username_final != username:
                    self.stdout.write(self.style.WARNING(
                        f'    Username {username} ya existe, usando {username_final}'
                    ))

                # Crear usuario
                user = User.objects.create_user(
                    username=username_final,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    email=f'{username_final}@esperanza.com.mx'
                )

                # Crear empleado
                empleado = Empleado.objects.create(
                    user=user,
                    codigo_empleado=codigo,
                    departamento=departamento,
                    puesto=puesto,
                    activo=True
                )

                self.stdout.write(self.style.SUCCESS(f'    Creado: {empleado}'))
                creados += 1

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'    Error: {e}'))
                errores += 1

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'Empleados creados: {creados}'))
        if update:
            self.stdout.write(self.style.SUCCESS(f'Empleados actualizados: {actualizados}'))
        self.stdout.write(self.style.WARNING(f'Ya existentes (omitidos): {existentes}'))
        if errores:
            self.stdout.write(self.style.ERROR(f'Errores: {errores}'))

    def generar_username(self, nombre_completo):
        """Genera un username a partir del nombre completo."""
        # Normalizar: quitar acentos
        nombre = unicodedata.normalize('NFD', nombre_completo)
        nombre = ''.join(c for c in nombre if unicodedata.category(c) != 'Mn')

        # Convertir a minúsculas
        nombre = nombre.lower()

        # Separar en partes
        partes = nombre.split()

        if len(partes) >= 2:
            # Usar primera letra del nombre + primer apellido
            username = partes[0][0] + partes[1]
        else:
            username = partes[0]

        # Quitar caracteres especiales
        username = re.sub(r'[^a-z0-9]', '', username)

        # Limitar longitud
        username = username[:20]

        return username
