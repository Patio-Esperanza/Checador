"""
Servicio para generar reportes en Excel
"""
from datetime import datetime, timedelta
from io import BytesIO
from django.db.models import Count, Q, Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from empleados.models import Empleado
from registros.models import RegistroAsistencia


class ExcelReportService:
    """Servicio para generar reportes de asistencia en Excel"""
    
    # Colores
    COLOR_HEADER = 'FF4472C4'
    COLOR_TOP_RETARDOS = 'FFFFC000'
    COLOR_FALTAS = 'FFFF0000'
    
    def __init__(self, fecha_inicio, fecha_fin):
        """
        Inicializa el servicio
        
        Args:
            fecha_inicio: Fecha de inicio del periodo
            fecha_fin: Fecha fin del periodo
        """
        self.fecha_inicio = fecha_inicio
        self.fecha_fin = fecha_fin
        self.workbook = Workbook()
        
    def generar_reporte_completo(self):
        """
        Genera el reporte completo con dos hojas:
        1. Concentrado: resumen por empleado
        2. Detalle: todos los registros
        
        Returns:
            BytesIO: Archivo Excel en memoria
        """
        # Eliminar hoja por defecto
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Crear hojas
        self._crear_hoja_concentrado()
        self._crear_hoja_detalle()
        
        # Guardar en memoria
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output
    
    def _crear_hoja_concentrado(self):
        """Crea la hoja de concentrado con resumen por empleado"""
        ws = self.workbook.create_sheet('Concentrado', 0)
        
        # Título del reporte
        ws.merge_cells('A1:F1')
        titulo_cell = ws['A1']
        titulo_cell.value = f'Reporte de Asistencias - {self.fecha_inicio.strftime("%d/%m/%Y")} al {self.fecha_fin.strftime("%d/%m/%Y")}'
        titulo_cell.font = Font(size=14, bold=True)
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Encabezados
        headers = ['Código', 'Nombre', 'Días Trabajados', 'Faltas', 'Retardos', 'Horas Totales']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos de empleados
        datos_empleados = self._obtener_datos_concentrado()
        
        # Escribir datos
        row = 4
        for empleado_data in datos_empleados:
            ws.cell(row, 1, empleado_data['codigo'])
            ws.cell(row, 2, empleado_data['nombre'])
            ws.cell(row, 3, empleado_data['dias_trabajados'])
            ws.cell(row, 4, empleado_data['faltas'])
            
            # Retardos con formato especial si están en el top 5
            retardos_cell = ws.cell(row, 5, empleado_data['retardos'])
            if empleado_data.get('top_retardos', False):
                retardos_cell.fill = PatternFill(start_color=self.COLOR_TOP_RETARDOS, end_color=self.COLOR_TOP_RETARDOS, fill_type='solid')
                retardos_cell.font = Font(bold=True)
            
            ws.cell(row, 6, round(empleado_data['horas_totales'], 2))
            row += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        
        # Aplicar bordes
        self._aplicar_bordes(ws, 3, len(datos_empleados) + 3, len(headers))
        
    def _crear_hoja_detalle(self):
        """Crea la hoja de detalle con todos los registros"""
        ws = self.workbook.create_sheet('Detalle de Registros')
        
        # Título
        ws.merge_cells('A1:H1')
        titulo_cell = ws['A1']
        titulo_cell.value = f'Detalle de Registros - {self.fecha_inicio.strftime("%d/%m/%Y")} al {self.fecha_fin.strftime("%d/%m/%Y")}'
        titulo_cell.font = Font(size=14, bold=True)
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Encabezados
        headers = ['Código', 'Nombre', 'Fecha', 'Entrada', 'Salida', 'Horas', 'Retardo', 'Notas']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener registros
        registros = RegistroAsistencia.objects.filter(
            fecha__gte=self.fecha_inicio,
            fecha__lte=self.fecha_fin
        ).select_related('empleado', 'empleado__user').order_by('empleado__codigo_empleado', 'fecha')
        
        # Escribir datos
        row = 4
        for registro in registros:
            ws.cell(row, 1, registro.empleado.codigo_empleado)
            ws.cell(row, 2, registro.empleado.nombre_completo)
            ws.cell(row, 3, registro.fecha.strftime('%d/%m/%Y'))
            ws.cell(row, 4, registro.hora_entrada.strftime('%H:%M') if registro.hora_entrada else '-')
            ws.cell(row, 5, registro.hora_salida.strftime('%H:%M') if registro.hora_salida else '-')
            ws.cell(row, 6, round(registro.horas_trabajadas, 2))
            
            retardo_cell = ws.cell(row, 7, 'Sí' if registro.retardo else 'No')
            if registro.retardo:
                retardo_cell.fill = PatternFill(start_color=self.COLOR_TOP_RETARDOS, end_color=self.COLOR_TOP_RETARDOS, fill_type='solid')
                retardo_cell.font = Font(bold=True)
            
            ws.cell(row, 8, registro.notas or '')
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 30
        
        # Aplicar bordes
        if row > 4:
            self._aplicar_bordes(ws, 3, row - 1, len(headers))
    
    def _obtener_datos_concentrado(self):
        """
        Obtiene los datos concentrados por empleado
        
        Returns:
            list: Lista de diccionarios con datos de cada empleado
        """
        # Obtener empleados activos que tienen registros en el periodo
        empleados = Empleado.objects.filter(
            activo=True,
            registros__fecha__gte=self.fecha_inicio,
            registros__fecha__lte=self.fecha_fin
        ).distinct()
        
        datos = []
        total_dias = (self.fecha_fin - self.fecha_inicio).days + 1
        
        for empleado in empleados:
            # Registros del empleado en el periodo
            registros = RegistroAsistencia.objects.filter(
                empleado=empleado,
                fecha__gte=self.fecha_inicio,
                fecha__lte=self.fecha_fin
            )
            
            dias_trabajados = registros.count()
            retardos = registros.filter(retardo=True).count()
            horas_totales = registros.aggregate(Sum('horas_trabajadas'))['horas_trabajadas__sum'] or 0
            
            # Calcular faltas (días sin registro - simplificado)
            # En un sistema real, considerarías días laborables, vacaciones, etc.
            faltas = max(0, total_dias - dias_trabajados)
            
            datos.append({
                'codigo': empleado.codigo_empleado,
                'nombre': empleado.nombre_completo,
                'dias_trabajados': dias_trabajados,
                'faltas': faltas,
                'retardos': retardos,
                'horas_totales': horas_totales,
                'empleado_id': empleado.id
            })
        
        # Ordenar por retardos descendente y marcar top 5
        datos.sort(key=lambda x: x['retardos'], reverse=True)
        for i in range(min(5, len(datos))):
            if datos[i]['retardos'] > 0:
                datos[i]['top_retardos'] = True
        
        # Reordenar por código de empleado
        datos.sort(key=lambda x: x['codigo'])
        
        return datos
    
    def _aplicar_bordes(self, ws, start_row, end_row, num_cols):
        """Aplica bordes a un rango de celdas"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                ws.cell(row, col).border = thin_border
    
    def obtener_top_retardos(self, limit=5):
        """
        Obtiene el top de empleados con más retardos
        
        Args:
            limit: Número de empleados a retornar
            
        Returns:
            list: Lista de diccionarios con código, nombre y cantidad de retardos
        """
        empleados = Empleado.objects.filter(
            activo=True,
            registros__fecha__gte=self.fecha_inicio,
            registros__fecha__lte=self.fecha_fin,
            registros__retardo=True
        ).annotate(
            total_retardos=Count('registros', filter=Q(registros__retardo=True))
        ).order_by('-total_retardos')[:limit]
        
        return [
            {
                'codigo': emp.codigo_empleado,
                'nombre': emp.nombre_completo,
                'retardos': emp.total_retardos
            }
            for emp in empleados
        ]
    
    def obtener_empleados_con_faltas(self):
        """
        Obtiene empleados que faltaron en el periodo
        
        Returns:
            list: Lista de diccionarios con código, nombre y fechas faltantes
        """
        # Simplificado: empleados activos que no tienen registros completos
        empleados = Empleado.objects.filter(activo=True)
        empleados_faltas = []
        
        for empleado in empleados:
            registros = RegistroAsistencia.objects.filter(
                empleado=empleado,
                fecha__gte=self.fecha_inicio,
                fecha__lte=self.fecha_fin
            ).count()
            
            # Si tiene menos registros que días en el periodo
            total_dias = (self.fecha_fin - self.fecha_inicio).days + 1
            if registros < total_dias:
                faltas = total_dias - registros
                empleados_faltas.append({
                    'codigo': empleado.codigo_empleado,
                    'nombre': empleado.nombre_completo,
                    'faltas': faltas
                })
        
        return empleados_faltas
